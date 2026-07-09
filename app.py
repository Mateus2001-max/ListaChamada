import streamlit as st
import pandas as pd
import sqlite3
import io
import os
from datetime import date, timedelta
from Crud.GerenciadorAlunos import GerenciadorAlunos
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from menu import gerar_relatorio_historico
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
import json

# 🔹 Configuração do Google Drive
FOLDER_ID = "1kNMGdts9a9gCKY8zQ909_pQCNW-YR3yj"

def autenticar_drive():
    credentials = Credentials.from_service_account_info(
        st.secrets["SERVICE_ACCOUNT"],  # já é dict
        scopes=["https://www.googleapis.com/auth/drive"]
    )
    service = build("drive", "v3", credentials=credentials)
    return service

def upload_to_drive(file_path, folder_id=FOLDER_ID):
    service = autenticar_drive()

    file_metadata = {
        "name": os.path.basename(file_path),
        "parents": [folder_id]
    }

    media = MediaFileUpload(
        file_path,
        resumable=True
    )

    arquivo = service.files().create(
        body=file_metadata,
        media_body=media,
        fields="id,name"
    ).execute()

    print(f"✅ Arquivo enviado: {arquivo['name']}")
    
def arquivo_existe_no_drive(nome_arquivo, folder_id=FOLDER_ID):
    service = autenticar_drive()

    query = (
        f"name='{nome_arquivo}' and "
        f"'{folder_id}' in parents and "
        f"trashed=false"
    )

    resultado = service.files().list(
        q=query,
        fields="files(id,name)"
    ).execute()

    return len(resultado.get("files", [])) > 0

# 🔹 Função principal de presença
def lista_presenca(alunos):
    st.title("📋 Lista de Presença")

    hoje = date.today().isoformat()
    conn = sqlite3.connect("escola.db")
    cursor = conn.cursor()
    cursor.execute("SELECT aluno_id, presente FROM presencas WHERE data = ?", (hoje,))
    presencas = dict(cursor.fetchall())
    conn.close()

    # Montar DataFrame
    dados = []
    for i, aluno in enumerate(alunos, start=1):
        dados.append({
            "ID": i,
            "PRESENTE": bool(presencas.get(aluno.id, 0)),
            "NOME": aluno.nome,
            "CIM Nº": aluno.rg,
            "GRAU": aluno.turma
        })
    df = pd.DataFrame(dados)

    # Editor interativo
    tabela_editada = st.data_editor(
        df,
        width="stretch",
        height=600,
        num_rows="dynamic",
        key="tabela_presenca"
    )

    # Botão para salvar alterações
    if st.button("💾 Salvar Alterações"):
        conn = sqlite3.connect("escola.db")
        cursor = conn.cursor()
        for _, linha in tabela_editada.iterrows():
            cursor.execute("""
                INSERT INTO alunos (id, nome, rg, turma)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(id) DO UPDATE SET nome=excluded.nome, rg=excluded.rg, turma=excluded.turma
            """, (linha["ID"], linha["NOME"], linha["CIM Nº"], linha["GRAU"]))

            presente = 1 if linha["PRESENTE"] else 0
            cursor.execute("""
                INSERT INTO presencas (aluno_id, data, presente)
                VALUES (?, ?, ?)
                ON CONFLICT(aluno_id, data) DO UPDATE SET presente=excluded.presente
            """, (linha["ID"], hoje, presente))
        conn.commit()
        conn.close()
        st.success("✅ Alterações salvas com sucesso!")

    # Botão para excluir alunos removidos
    if st.button("🗑️ Remover alunos excluídos"):
        conn = sqlite3.connect("escola.db")
        cursor = conn.cursor()
        ids_atuais = set(tabela_editada["ID"].dropna().astype(int))
        cursor.execute("SELECT id FROM alunos")
        ids_banco = set(row[0] for row in cursor.fetchall())
        ids_removidos = ids_banco - ids_atuais
        for id_removido in ids_removidos:
            cursor.execute("DELETE FROM alunos WHERE id=?", (id_removido,))
        conn.commit()
        conn.close()
        st.success("✅ Alunos removidos do banco com sucesso!")
        st.rerun()

    # Estatísticas
    total = len(tabela_editada)
    presentes = sum(bool(x) for x in tabela_editada["PRESENTE"])
    porcentagem = (presentes / total * 100) if total > 0 else 0

    st.write(f"**Total de alunos:** {total}")
    st.write(f"**Presentes:** {presentes}")
    st.progress(porcentagem / 100)
    st.success(f"**Porcentagem de presença:** {porcentagem:.2f}%")

    # Criar DataFrame para download
    df_download = tabela_editada.copy()
    df_download["PRESENTE"] = df_download["PRESENTE"].apply(lambda x: "Sim" if x else "Não")
    df_download = df_download[["ID", "PRESENTE", "NOME", "CIM Nº", "GRAU"]]

    # Linha de resumo
    resumo = pd.DataFrame([{
        "ID": "",
        "PRESENTE": f"{presentes}/{total} ({porcentagem:.2f}%)",
        "NOME": "Resumo",
        "CIM Nº": "",
        "GRAU": ""
    }])
    df_final = pd.concat([df_download, resumo], ignore_index=True)

    # Converter para Excel em memória
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_final.to_excel(writer, index=False, sheet_name="Presenca", startrow=6)
        ws = writer.book["Presenca"]

        # Inserir logotipo
        logo = Image("logo.png")
        logo.width = 120
        logo.height = 120
        ws.add_image(logo, "A1")

        # Cabeçalho
        ws.merge_cells("C1:E1")
        ws.merge_cells("C2:E2")
        ws.merge_cells("C3:E3")
        ws.merge_cells("C4:E4")

        ws["C1"] = "A...G...D...G...A...D...U..."
        ws["C2"] = "ESP... LOJ... SIMB... TERCEIRO MILÊNIO Nº 2.825"
        ws["C3"] = "∴A∴A∴A∴ REUNIÕES 5.ª FEIRA ÀS 19:30 H."
        ws["C4"] = "FEDERADA AO G∴O∴B∴ E JURISDICIONADA AO G∴O∴B∴M∴S∴"
        ws["C6"] = f"Data: {date.today().strftime('%d/%m/%Y')}"
        ws["C6"].alignment = Alignment(horizontal="center", vertical="center")
        ws["C6"].font = Font(bold=True, size=11)

        for row in range(1, 5):
            cell = ws[f"C{row}"]
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True, size=12)

        # Bordas
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=Side(style="thin"), bottom=Side(style="thin"))

        for row in ws.iter_rows(min_row=7, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Ajustar largura
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

    # Botão para baixar
    st.download_button(
        label="📥 Baixar presença de hoje",
        data=buffer.getvalue(),
        file_name=f"presenca_{date.today().isoformat()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# 🔹 Automático: cria e envia relatório de ontem
ontem = (date.today() - timedelta(days=1)).isoformat()
arquivo_excel = f"presenca_{ontem}.xlsx"
caminho_local = f"Relatorios/{arquivo_excel}"

if not os.path.exists(caminho_local):
    gerar_relatorio_historico()
    upload_to_drive(caminho_local)
    st.success(f"📂 Relatório de ontem criado e enviado para o Google Drive: {arquivo_excel}")
else:
    if arquivo_existe_no_drive(arquivo_excel):
        st.info(f"📂 Relatório de ontem já está no Google Drive: {arquivo_excel}")
    else:
        upload_to_drive(caminho_local)
        st.success(f"📂 Relatório de ontem enviado para o Google Drive: {arquivo_excel}")

# Chamada da função
ger = GerenciadorAlunos()
alunos = ger.listar()

if alunos:
    lista_presenca(alunos)
else:
    st.warning("Nenhum aluno cadastrado no banco de dados.")

