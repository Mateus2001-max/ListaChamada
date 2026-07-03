from Crud.GerenciadorAlunos import GerenciadorAlunos
#from Interface.Interface_tkinter import  abrir_lista_presenca
from datetime import datetime
from docx import Document
from docx.shared import Pt
from docx.oxml.ns import qn
from docx.oxml import OxmlElement
import os
import pandas as pd
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
import sqlite3
from datetime import date, timedelta
def exibir_menu():
    gerar_relatorio_historico()
    
    ger = GerenciadorAlunos()

    while True:
        print("\n===== MENU ALUNOS =====")
        print("1 - Cadastrar aluno")
        print("2 - Listar alunos")
        print("3 - Atualizar aluno")
        print("4 - Remover aluno")
        print("5 - Lista de Presença")
        print("0 - Sair")

        opcao = input("Escolha uma opção: ")

        if opcao == "1":
            nome = input("Digite o nome do aluno: ")
            rg = input("Digite o RG do aluno: ")
            turma = input("Digite a turma do aluno (ex: 3º ano E.M): ")
            ger.adicionar(nome, rg, turma)
            print("Aluno cadastrado com sucesso!")


        elif opcao == "2":
            alunos = ger.listar()
            if alunos:
                for aluno in alunos:
                    print(aluno)
            else:
                print("Nenhum aluno cadastrado.")

        elif opcao == "3":
            id = int(input("Digite o ID do aluno a atualizar: "))
            novo_nome = input("Digite o novo nome: ")
            novo_rg = input("Digite o novo RG: ")
            nova_turma = input("Digite a nova turma: ")
            ger.atualizar(id, novo_nome, novo_rg, nova_turma)



        elif opcao == "4":
            try:
                id = int(input("Digite o ID do aluno a remover: "))
                # checar se o ID existe antes de remover
                alunos = ger.listar()
                ids_existentes = [aluno.id for aluno in alunos]

                if id not in ids_existentes:
                    print(f"⚠️ Nenhum aluno encontrado com ID {id}.")
                else:
                    ger.remover(id)
            except ValueError:
                print("⚠️ O ID deve ser um número válido!")

        elif opcao == "5":
            alunos = ger.listar()
            if alunos:
                #abrir_lista_presenca(alunos)
            else:
                print("Nenhum aluno cadastrado.")



        elif opcao == "0":
            print("Saindo do sistema...")
            break

        else:
            print("Opção inválida, tente novamente.")

def gerar_relatorio_historico():
    # Criar pasta se não existir
    os.makedirs("Relatorios", exist_ok=True)

    # Data de ontem
    ontem = (date.today() - timedelta(days=1)).isoformat()
    arquivo_excel = f"Relatorios/presenca_{ontem}.xlsx"

    # Buscar alunos
    ger = GerenciadorAlunos()
    alunos = ger.listar()

    # Conectar ao banco para pegar presenças de ontem
    conn = sqlite3.connect("escola.db")
    cursor = conn.cursor()
    cursor.execute("SELECT aluno_id, presente FROM presencas WHERE data = ?", (ontem,))
    presencas = dict(cursor.fetchall())
    conn.close()

    # Montar DataFrame
    dados = []
    for i, aluno in enumerate(alunos, start=1):
        dados.append({
            "ID": i,
            "PRESENTE": "Sim" if presencas.get(aluno.id, 0) == 1 else "Não",
            "NOME": aluno.nome,
            "CIM Nº": aluno.rg,
            "GRAU": aluno.turma
        })
    df = pd.DataFrame(dados)

    # Estatísticas
    total = len(df)
    presentes = sum(df["PRESENTE"] == "Sim")
    porcentagem = (presentes / total * 100) if total > 0 else 0

    resumo = pd.DataFrame([{
        "ID": "",
        "PRESENTE": f"{presentes}/{total} ({porcentagem:.2f}%)",
        "NOME": "Resumo",
        "CIM Nº": "",
        "GRAU": ""
    }])
    df_final = pd.concat([df, resumo], ignore_index=True)

    # Criar Excel com formatação e cabeçalho
    with pd.ExcelWriter(arquivo_excel, engine="openpyxl") as writer:
        df_final.to_excel(writer, index=False, sheet_name="Presenca", startrow=6)
        ws = writer.book["Presenca"]

        # Inserir logotipo
        logo = Image("logo.png")
        logo.width = 120
        logo.height = 120
        ws.add_image(logo, "A1")

        # Cabeçalho de texto
        ws.merge_cells("C1:E1")
        ws.merge_cells("C2:E2")
        ws.merge_cells("C3:E3")
        ws.merge_cells("C4:E4")

        ws["C1"] = "A...G...D...G...A...D...U..."
        ws["C2"] = "ESP... LOJ... SIMB... TERCEIRO MILÊNIO Nº 2.825"
        ws["C3"] = "∴A∴A∴A∴ REUNIÕES 5.ª FEIRA ÀS 19:30 H."
        ws["C4"] = "FEDERADA AO G∴O∴B∴ E JURISDICIONADA AO G∴O∴B∴M∴S∴"
        ws["C6"] = f"Data: {ontem}"
        ws["C6"].alignment = Alignment(horizontal="center", vertical="center")
        ws["C6"].font = Font(bold=True, size=11)

        for row in range(1, 5):
            cell = ws[f"C{row}"]
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True, size=12)

        # Bordas finas
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=Side(style="thin"), bottom=Side(style="thin"))

        for row in ws.iter_rows(min_row=7, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Ajustar largura automática
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

    print(f"✅ Relatório histórico criado: {arquivo_excel}")


